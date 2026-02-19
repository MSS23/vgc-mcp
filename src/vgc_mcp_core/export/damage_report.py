"""Generate damage calculation reports in Excel and PDF formats."""

import os
import tempfile
from typing import Optional

from ..calc.bulk_calc import BulkCalcSummary, get_results_for_scenario  # noqa: I001
from ..calc.damage import format_percent

# =============================================================================
# Color coding for KO verdicts
# =============================================================================

# (R, G, B) tuples for cell backgrounds
OHKO_COLOR = (198, 239, 206)       # Green
TWOHKO_COLOR = (255, 235, 156)     # Yellow
THREEHKO_COLOR = (252, 213, 180)   # Orange
FOURHKO_PLUS_COLOR = (255, 199, 206)  # Red/Pink
HEADER_COLOR = (68, 114, 196)      # Blue header
HEADER_FONT_COLOR = (255, 255, 255)  # White text


def _get_ko_color(ko_chance: str) -> tuple[int, int, int]:
    """Map KO verdict to a color tuple."""
    ko_lower = ko_chance.lower()
    if "ohko" in ko_lower or "1hko" in ko_lower:
        return OHKO_COLOR
    if "2hko" in ko_lower:
        return TWOHKO_COLOR
    if "3hko" in ko_lower:
        return THREEHKO_COLOR
    return FOURHKO_PLUS_COLOR


def _format_cell_value(min_pct: float, max_pct: float, ko_chance: str) -> str:
    """Format a damage result as a cell value."""
    min_str = format_percent(min_pct)
    max_str = format_percent(max_pct)
    return f"{min_str}-{max_str}% ({ko_chance})"


# =============================================================================
# Excel report generation
# =============================================================================

def generate_excel_report(
    summary: BulkCalcSummary,
    output_path: Optional[str] = None,
) -> str:
    """Generate an Excel report with color-coded damage results.

    Creates one sheet per scenario. Each sheet has:
    - Header row with attacker info
    - Columns: Defender | Their Spread | Item | Move1 | Move2 | ...
    - Color-coded cells based on KO verdict

    Args:
        summary: BulkCalcSummary from run_bulk_calcs
        output_path: Optional output file path. If None, creates a temp file.

    Returns:
        Path to the generated .xlsx file
    """
    try:
        from openpyxl import Workbook  # noqa: I001
        from openpyxl.styles import (
            Alignment, Border, Font, PatternFill, Side,
        )
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. Install with: pip install openpyxl"
        )

    if output_path is None:
        output_path = os.path.join(
            tempfile.gettempdir(),
            f"{summary.attacker_name}_damage_report.xlsx"
        )

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=13)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Get unique defender names in order
    seen_defenders = []
    for r in summary.results:
        if r.defender_name not in seen_defenders:
            seen_defenders.append(r.defender_name)

    for scenario_name in summary.scenario_names:
        scenario_results = get_results_for_scenario(summary, scenario_name)
        if not scenario_results:
            continue

        # Find display name from first result
        display_name = scenario_results[0].scenario_display
        sheet_name = display_name[:31]  # Excel max sheet name length
        ws = wb.create_sheet(title=sheet_name)

        # Title row - attacker info
        end_col = 3 + len(summary.move_names)
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=end_col,
        )
        title_cell = ws.cell(row=1, column=1)
        atk_title = summary.attacker_name.replace('-', ' ').title()
        title_cell.value = f"{atk_title} — {summary.attacker_spread_str}"
        title_cell.font = title_font

        # Scenario label
        ws.merge_cells(
            start_row=2, start_column=1,
            end_row=2, end_column=end_col,
        )
        scenario_cell = ws.cell(row=2, column=1)
        scenario_cell.value = f"Scenario: {display_name}"
        scenario_cell.font = Font(italic=True, size=11)

        # Header row
        header_row = 4
        headers = ["Defender", "Spread", "Item"] + [
            m.replace("-", " ").title() for m in summary.move_names
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        data_row = header_row + 1
        for defender_name in seen_defenders:
            ws.cell(row=data_row, column=1).value = defender_name.replace("-", " ").title()
            ws.cell(row=data_row, column=1).border = thin_border
            ws.cell(row=data_row, column=1).font = Font(bold=True)

            spread = summary.defender_spreads.get(defender_name, "")
            ws.cell(row=data_row, column=2).value = spread
            ws.cell(row=data_row, column=2).border = thin_border

            item = summary.defender_items.get(defender_name, "")
            item_val = item.replace("-", " ").title() if item != "None" else ""
            ws.cell(row=data_row, column=3).value = item_val
            ws.cell(row=data_row, column=3).border = thin_border

            # Move columns
            for move_idx, move_name in enumerate(summary.move_names):
                col = 4 + move_idx
                # Find the result for this defender + move + scenario
                matching = [
                    r for r in scenario_results
                    if r.defender_name == defender_name and r.move_name == move_name
                ]
                if matching:
                    r = matching[0]
                    cell = ws.cell(row=data_row, column=col)
                    cell.value = _format_cell_value(r.min_pct, r.max_pct, r.ko_chance)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

                    # Color code
                    color = _get_ko_color(r.ko_chance)
                    hex_color = f"{color[0]:02X}{color[1]:02X}{color[2]:02X}"
                    cell.fill = PatternFill(
                        start_color=hex_color, end_color=hex_color,
                        fill_type="solid",
                    )
                else:
                    ws.cell(row=data_row, column=col).value = "—"
                    ws.cell(row=data_row, column=col).border = thin_border

            data_row += 1

        # Auto-size columns (approximate)
        for col_idx in range(1, 4 + len(summary.move_names)):
            ws.column_dimensions[ws.cell(row=header_row, column=col_idx).column_letter].width = 22

        # Freeze header
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    wb.save(output_path)
    return output_path


# =============================================================================
# PDF report generation
# =============================================================================

def generate_pdf_report(
    summary: BulkCalcSummary,
    output_path: Optional[str] = None,
) -> str:
    """Generate a PDF report with color-coded damage results.

    Args:
        summary: BulkCalcSummary from run_bulk_calcs
        output_path: Optional output file path. If None, creates a temp file.

    Returns:
        Path to the generated .pdf file
    """
    try:
        from fpdf import FPDF
    except ImportError:
        raise ImportError(
            "fpdf2 is required for PDF export. Install with: pip install fpdf2"
        )

    if output_path is None:
        output_path = os.path.join(
            tempfile.gettempdir(),
            f"{summary.attacker_name}_damage_report.pdf"
        )

    pdf = FPDF(orientation="L", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)

    # Get unique defender names in order
    seen_defenders = []
    for r in summary.results:
        if r.defender_name not in seen_defenders:
            seen_defenders.append(r.defender_name)

    num_moves = len(summary.move_names)
    # Calculate column widths for landscape A4 (297mm usable ~277mm)
    name_width = 35
    spread_width = 55
    item_width = 30
    remaining = 277 - name_width - spread_width - item_width
    move_width = remaining / max(num_moves, 1)

    for scenario_name in summary.scenario_names:
        scenario_results = get_results_for_scenario(summary, scenario_name)
        if not scenario_results:
            continue

        display_name = scenario_results[0].scenario_display

        pdf.add_page()

        # Title
        pdf.set_font("Helvetica", "B", 14)
        attacker_title = summary.attacker_name.replace("-", " ").title()
        title_text = f"{attacker_title} - {summary.attacker_spread_str}"
        pdf.cell(
            0, 10, title_text, new_x="LMARGIN", new_y="NEXT",
        )

        # Scenario subtitle
        pdf.set_font("Helvetica", "I", 11)
        pdf.cell(0, 8, f"Scenario: {display_name}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        # Header row
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(*HEADER_COLOR)
        pdf.set_text_color(*HEADER_FONT_COLOR)

        pdf.cell(name_width, 8, "Defender", border=1, fill=True, align="C")
        pdf.cell(spread_width, 8, "Spread", border=1, fill=True, align="C")
        pdf.cell(item_width, 8, "Item", border=1, fill=True, align="C")
        for move_name in summary.move_names:
            move_label = move_name.replace("-", " ").title()[:15]
            pdf.cell(
                move_width, 8, move_label,
                border=1, fill=True, align="C",
            )
        pdf.ln()

        # Data rows
        pdf.set_font("Helvetica", "", 7)
        pdf.set_text_color(0, 0, 0)

        for defender_name in seen_defenders:
            pdf.cell(name_width, 7, defender_name.replace("-", " ").title()[:20], border=1)

            spread = summary.defender_spreads.get(defender_name, "")
            pdf.cell(spread_width, 7, spread[:30], border=1)

            item = summary.defender_items.get(defender_name, "")
            item_display = item.replace("-", " ").title() if item != "None" else ""
            pdf.cell(item_width, 7, item_display[:15], border=1)

            for move_name in summary.move_names:
                matching = [
                    r for r in scenario_results
                    if r.defender_name == defender_name and r.move_name == move_name
                ]
                if matching:
                    r = matching[0]
                    cell_text = _format_cell_value(r.min_pct, r.max_pct, r.ko_chance)
                    color = _get_ko_color(r.ko_chance)
                    pdf.set_fill_color(*color)
                    pdf.cell(move_width, 7, cell_text[:25], border=1, fill=True, align="C")
                else:
                    pdf.cell(move_width, 7, "—", border=1, align="C")

            pdf.ln()

    pdf.output(output_path)
    return output_path
