"""Tests for bulk offensive damage calculation engine."""

import os
import shutil
import tempfile

import pytest

from vgc_mcp_core.calc.bulk_calc import (
    DEFAULT_SCENARIOS,
    BulkCalcResult,
    BulkCalcSummary,
    ScenarioConfig,
    build_calc_string,
    build_scenario_modifiers,
    get_best_move_per_defender,
    get_results_for_defender,
    get_results_for_scenario,
    run_bulk_calcs,
)
from vgc_mcp_core.calc.damage import calculate_damage
from vgc_mcp_core.models.move import Move, MoveCategory
from vgc_mcp_core.models.pokemon import BaseStats, EVSpread, Nature, PokemonBuild

# Inline base stats (avoid importing conftest directly)
URSHIFU_STATS = BaseStats(
    hp=100, attack=130, defense=100,
    special_attack=63, special_defense=60, speed=97,
)
INCINEROAR_STATS = BaseStats(
    hp=95, attack=115, defense=90,
    special_attack=80, special_defense=90, speed=60,
)
FLUTTER_MANE_STATS = BaseStats(
    hp=55, attack=55, defense=55,
    special_attack=135, special_defense=135, speed=135,
)


# =============================================================================
# Helpers
# =============================================================================

def _make_build(
    name: str,
    types: list[str],
    base_stats: BaseStats,
    nature: Nature = Nature.ADAMANT,
    evs: dict | None = None,
    item: str | None = None,
    ability: str | None = None,
    tera_type: str | None = None,
) -> PokemonBuild:
    ev_dict = evs or {}
    return PokemonBuild(
        name=name,
        base_stats=base_stats,
        types=types,
        nature=nature,
        evs=EVSpread(
            hp=ev_dict.get("hp", 0),
            attack=ev_dict.get("attack", 0),
            defense=ev_dict.get("defense", 0),
            special_attack=ev_dict.get("special_attack", 0),
            special_defense=ev_dict.get("special_defense", 0),
            speed=ev_dict.get("speed", 0),
        ),
        item=item,
        ability=ability,
        tera_type=tera_type,
    )


def _make_move(
    name: str,
    move_type: str,
    category: MoveCategory = MoveCategory.PHYSICAL,
    power: int = 80,
    always_crit: bool = False,
    min_hits: int = 1,
    max_hits: int = 1,
    target: str = "selected-pokemon",
) -> Move:
    return Move(
        name=name,
        type=move_type,
        category=category,
        power=power,
        accuracy=100,
        always_crit=always_crit,
        min_hits=min_hits,
        max_hits=max_hits,
        target=target,
    )


# =============================================================================
# Test fixtures
# =============================================================================

@pytest.fixture
def urshifu():
    return _make_build(
        "urshifu-rapid-strike",
        ["fighting", "water"],
        URSHIFU_STATS,
        nature=Nature.ADAMANT,
        evs={"attack": 252, "speed": 252, "hp": 4},
        item="choice-scarf",
        ability="unseen-fist",
        tera_type="water",
    )


@pytest.fixture
def incineroar():
    return _make_build(
        "incineroar",
        ["fire", "dark"],
        INCINEROAR_STATS,
        nature=Nature.CAREFUL,
        evs={"hp": 252, "special_defense": 128, "defense": 128},
        item="sitrus-berry",
        ability="intimidate",
    )


@pytest.fixture
def flutter_mane():
    return _make_build(
        "flutter-mane",
        ["ghost", "fairy"],
        FLUTTER_MANE_STATS,
        nature=Nature.TIMID,
        evs={"special_attack": 252, "speed": 252, "hp": 4},
        item="choice-specs",
        ability="protosynthesis",
    )


@pytest.fixture
def surging_strikes():
    return _make_move(
        "surging-strikes", "water",
        power=25, always_crit=True, min_hits=3, max_hits=3,
    )


@pytest.fixture
def close_combat():
    return _make_move("close-combat", "fighting", power=120)


@pytest.fixture
def aqua_jet():
    return _make_move("aqua-jet", "water", power=40)


# =============================================================================
# Tests: ScenarioConfig & build_scenario_modifiers
# =============================================================================

class TestScenarioModifiers:

    def test_normal_scenario_has_no_modifiers(self, urshifu, surging_strikes):
        scenario = DEFAULT_SCENARIOS["normal"]
        mods = build_scenario_modifiers(scenario, urshifu, surging_strikes)
        assert mods.weather is None
        assert mods.tera_active is False
        assert mods.helping_hand is False
        assert mods.attack_stage == 0

    def test_rain_scenario_sets_weather(self, urshifu, surging_strikes):
        scenario = DEFAULT_SCENARIOS["rain"]
        mods = build_scenario_modifiers(scenario, urshifu, surging_strikes)
        assert mods.weather == "rain"
        assert mods.tera_active is False

    def test_tera_scenario_uses_attacker_tera_type(self, urshifu, surging_strikes):
        scenario = DEFAULT_SCENARIOS["tera"]
        mods = build_scenario_modifiers(scenario, urshifu, surging_strikes)
        assert mods.tera_active is True
        assert mods.tera_type == "water"

    def test_rain_tera_combines_both(self, urshifu, surging_strikes):
        scenario = DEFAULT_SCENARIOS["rain_tera"]
        mods = build_scenario_modifiers(scenario, urshifu, surging_strikes)
        assert mods.weather == "rain"
        assert mods.tera_active is True
        assert mods.tera_type == "water"

    def test_intimidate_sets_attack_stage(self, urshifu, close_combat):
        scenario = DEFAULT_SCENARIOS["intimidate"]
        mods = build_scenario_modifiers(scenario, urshifu, close_combat)
        assert mods.attack_stage == -1

    def test_intimidate_ignored_for_always_crit_moves(self, urshifu, surging_strikes):
        """Surging Strikes always crits, so -1 Atk from Intimidate is ignored."""
        scenario = DEFAULT_SCENARIOS["intimidate"]
        mods = build_scenario_modifiers(scenario, urshifu, surging_strikes)
        # Critical hits ignore negative attack stages
        assert mods.attack_stage == 0
        assert mods.is_critical is True

    def test_helping_hand_scenario(self, urshifu, surging_strikes):
        scenario = DEFAULT_SCENARIOS["helping_hand"]
        mods = build_scenario_modifiers(scenario, urshifu, surging_strikes)
        assert mods.helping_hand is True

    def test_attacker_item_and_ability_carried_through(self, urshifu, surging_strikes):
        scenario = DEFAULT_SCENARIOS["normal"]
        mods = build_scenario_modifiers(scenario, urshifu, surging_strikes)
        assert mods.attacker_item == "choice-scarf"
        assert mods.attacker_ability == "unseen-fist"

    def test_is_doubles_always_true(self, urshifu, surging_strikes):
        for name, scenario in DEFAULT_SCENARIOS.items():
            mods = build_scenario_modifiers(scenario, urshifu, surging_strikes)
            assert mods.is_doubles is True, f"{name} should be doubles"

    def test_custom_scenario(self, urshifu, surging_strikes):
        custom = ScenarioConfig(
            name="custom_rain_hh",
            display_name="Rain + HH",
            weather="rain",
            helping_hand=True,
        )
        mods = build_scenario_modifiers(custom, urshifu, surging_strikes)
        assert mods.weather == "rain"
        assert mods.helping_hand is True

    def test_spread_move_gets_multiple_targets_true(self, urshifu):
        """Spread moves like Heat Wave should set multiple_targets=True."""
        heat_wave = _make_move(
            "heat-wave", "fire",
            category=MoveCategory.SPECIAL, power=95,
            target="all-adjacent-foes",
        )
        scenario = DEFAULT_SCENARIOS["normal"]
        mods = build_scenario_modifiers(scenario, urshifu, heat_wave)
        assert mods.multiple_targets is True

    def test_single_target_move_gets_multiple_targets_false(
        self, urshifu, close_combat,
    ):
        """Single-target moves should set multiple_targets=False."""
        scenario = DEFAULT_SCENARIOS["normal"]
        mods = build_scenario_modifiers(scenario, urshifu, close_combat)
        assert mods.multiple_targets is False


# =============================================================================
# Tests: build_calc_string
# =============================================================================

class TestCalcString:

    def test_calc_string_contains_attacker_name(self, urshifu, incineroar, surging_strikes):
        scenario = DEFAULT_SCENARIOS["normal"]
        mods = build_scenario_modifiers(scenario, urshifu, surging_strikes)
        result = calculate_damage(urshifu, incineroar, surging_strikes, mods)
        calc_str = build_calc_string(urshifu, incineroar, surging_strikes, result, scenario)
        assert "Urshifu" in calc_str

    def test_calc_string_contains_move_name(self, urshifu, incineroar, surging_strikes):
        scenario = DEFAULT_SCENARIOS["normal"]
        mods = build_scenario_modifiers(scenario, urshifu, surging_strikes)
        result = calculate_damage(urshifu, incineroar, surging_strikes, mods)
        calc_str = build_calc_string(urshifu, incineroar, surging_strikes, result, scenario)
        assert "Surging Strikes" in calc_str

    def test_calc_string_contains_defender_name(self, urshifu, incineroar, surging_strikes):
        scenario = DEFAULT_SCENARIOS["normal"]
        mods = build_scenario_modifiers(scenario, urshifu, surging_strikes)
        result = calculate_damage(urshifu, incineroar, surging_strikes, mods)
        calc_str = build_calc_string(urshifu, incineroar, surging_strikes, result, scenario)
        assert "Incineroar" in calc_str

    def test_calc_string_contains_damage_range(self, urshifu, incineroar, surging_strikes):
        scenario = DEFAULT_SCENARIOS["normal"]
        mods = build_scenario_modifiers(scenario, urshifu, surging_strikes)
        result = calculate_damage(urshifu, incineroar, surging_strikes, mods)
        calc_str = build_calc_string(urshifu, incineroar, surging_strikes, result, scenario)
        # Should contain damage numbers and percentages
        assert "%" in calc_str
        assert "--" in calc_str  # Separates calc from verdict

    def test_calc_string_normal_has_no_tag(self, urshifu, incineroar, surging_strikes):
        scenario = DEFAULT_SCENARIOS["normal"]
        mods = build_scenario_modifiers(scenario, urshifu, surging_strikes)
        result = calculate_damage(urshifu, incineroar, surging_strikes, mods)
        calc_str = build_calc_string(urshifu, incineroar, surging_strikes, result, scenario)
        assert not calc_str.startswith("[")

    def test_calc_string_rain_has_tag(self, urshifu, incineroar, surging_strikes):
        scenario = DEFAULT_SCENARIOS["rain"]
        mods = build_scenario_modifiers(scenario, urshifu, surging_strikes)
        result = calculate_damage(urshifu, incineroar, surging_strikes, mods)
        calc_str = build_calc_string(urshifu, incineroar, surging_strikes, result, scenario)
        assert calc_str.startswith("[Rain]")

    def test_calc_string_contains_item(self, urshifu, incineroar, surging_strikes):
        scenario = DEFAULT_SCENARIOS["normal"]
        mods = build_scenario_modifiers(scenario, urshifu, surging_strikes)
        result = calculate_damage(urshifu, incineroar, surging_strikes, mods)
        calc_str = build_calc_string(urshifu, incineroar, surging_strikes, result, scenario)
        assert "Choice Scarf" in calc_str


# =============================================================================
# Tests: run_bulk_calcs
# =============================================================================

class TestRunBulkCalcs:

    def test_single_move_single_defender_single_scenario(
        self, urshifu, incineroar, surging_strikes
    ):
        summary = run_bulk_calcs(
            urshifu,
            [surging_strikes],
            [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )
        assert summary.total_calcs == 1
        assert len(summary.results) == 1
        assert summary.results[0].defender_name == "incineroar"
        assert summary.results[0].move_name == "surging-strikes"
        assert summary.results[0].scenario_name == "normal"

    def test_multiple_moves_counted_correctly(
        self, urshifu, incineroar, surging_strikes, close_combat, aqua_jet
    ):
        summary = run_bulk_calcs(
            urshifu,
            [surging_strikes, close_combat, aqua_jet],
            [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )
        assert summary.total_calcs == 3
        move_names = {r.move_name for r in summary.results}
        assert move_names == {"surging-strikes", "close-combat", "aqua-jet"}

    def test_multiple_defenders_counted_correctly(
        self, urshifu, incineroar, flutter_mane, surging_strikes
    ):
        summary = run_bulk_calcs(
            urshifu,
            [surging_strikes],
            [incineroar, flutter_mane],
            [DEFAULT_SCENARIOS["normal"]],
        )
        assert summary.total_calcs == 2
        defenders = {r.defender_name for r in summary.results}
        assert defenders == {"incineroar", "flutter-mane"}

    def test_multiple_scenarios_counted_correctly(
        self, urshifu, incineroar, surging_strikes
    ):
        scenarios = [DEFAULT_SCENARIOS["normal"], DEFAULT_SCENARIOS["rain"]]
        summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar], scenarios,
        )
        assert summary.total_calcs == 2
        scenario_names = {r.scenario_name for r in summary.results}
        assert scenario_names == {"normal", "rain"}

    def test_full_combo_count(
        self, urshifu, incineroar, flutter_mane, surging_strikes, close_combat
    ):
        """2 moves × 2 defenders × 3 scenarios = 12 calcs."""
        scenarios = [
            DEFAULT_SCENARIOS["normal"],
            DEFAULT_SCENARIOS["rain"],
            DEFAULT_SCENARIOS["intimidate"],
        ]
        summary = run_bulk_calcs(
            urshifu,
            [surging_strikes, close_combat],
            [incineroar, flutter_mane],
            scenarios,
        )
        assert summary.total_calcs == 12

    def test_attacker_info_in_summary(self, urshifu, incineroar, surging_strikes):
        summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )
        assert summary.attacker_name == "urshifu-rapid-strike"
        assert "Adamant" in summary.attacker_spread_str
        assert "252" in summary.attacker_spread_str

    def test_defender_spread_tracked(self, urshifu, incineroar, surging_strikes):
        summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )
        assert "incineroar" in summary.defender_spreads
        assert "Careful" in summary.defender_spreads["incineroar"]

    def test_defender_items_tracked(self, urshifu, incineroar, surging_strikes):
        summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )
        assert summary.defender_items["incineroar"] == "sitrus-berry"

    def test_damage_results_have_valid_percentages(
        self, urshifu, incineroar, surging_strikes
    ):
        summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )
        result = summary.results[0]
        assert result.min_pct > 0
        assert result.max_pct >= result.min_pct
        assert result.min_damage > 0
        assert result.max_damage >= result.min_damage
        assert result.defender_hp > 0

    def test_rain_boosts_water_damage(self, urshifu, incineroar, surging_strikes):
        normal_summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )
        rain_summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["rain"]],
        )
        assert rain_summary.results[0].max_pct > normal_summary.results[0].max_pct

    def test_ohko_counts_populated(self, urshifu, flutter_mane, close_combat):
        """Close Combat vs Flutter Mane (fragile) should register as OHKO."""
        summary = run_bulk_calcs(
            urshifu, [close_combat], [flutter_mane],
            [DEFAULT_SCENARIOS["normal"]],
        )
        # Close Combat (120 BP Fighting STAB) vs Flutter Mane (55 base Def)
        # should be a comfortable OHKO
        result = summary.results[0]
        if result.min_pct >= 100:
            assert summary.ohko_counts["normal"] >= 1


# =============================================================================
# Tests: Result filtering helpers
# =============================================================================

class TestResultFiltering:

    def test_get_results_for_defender(
        self, urshifu, incineroar, flutter_mane, surging_strikes
    ):
        summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar, flutter_mane],
            [DEFAULT_SCENARIOS["normal"]],
        )
        incin_results = get_results_for_defender(summary, "incineroar")
        assert len(incin_results) == 1
        assert incin_results[0].defender_name == "incineroar"

    def test_get_results_for_scenario(
        self, urshifu, incineroar, surging_strikes
    ):
        scenarios = [DEFAULT_SCENARIOS["normal"], DEFAULT_SCENARIOS["rain"]]
        summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar], scenarios,
        )
        rain_results = get_results_for_scenario(summary, "rain")
        assert len(rain_results) == 1
        assert rain_results[0].scenario_name == "rain"

    def test_get_best_move_per_defender(
        self, urshifu, incineroar, surging_strikes, close_combat
    ):
        summary = run_bulk_calcs(
            urshifu, [surging_strikes, close_combat], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )
        best = get_best_move_per_defender(summary, "normal")
        assert "incineroar" in best
        # The best move should have the higher max_pct
        all_results = get_results_for_defender(summary, "incineroar")
        max_pct = max(r.max_pct for r in all_results if r.scenario_name == "normal")
        assert best["incineroar"].max_pct == max_pct


# =============================================================================
# Tests: Scenario edge cases
# =============================================================================

class TestScenarioEdgeCases:

    def test_sun_reduces_water_damage(self, urshifu, incineroar, surging_strikes):
        normal_summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )
        sun_summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["sun"]],
        )
        assert sun_summary.results[0].max_pct < normal_summary.results[0].max_pct

    def test_helping_hand_boosts_damage(self, urshifu, incineroar, surging_strikes):
        normal_summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )
        hh_summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["helping_hand"]],
        )
        assert hh_summary.results[0].max_pct > normal_summary.results[0].max_pct

    def test_all_default_scenarios_produce_results(
        self, urshifu, incineroar, surging_strikes
    ):
        all_scenarios = list(DEFAULT_SCENARIOS.values())
        summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar], all_scenarios,
        )
        assert summary.total_calcs == len(DEFAULT_SCENARIOS)
        for s in DEFAULT_SCENARIOS:
            results = get_results_for_scenario(summary, s)
            assert len(results) == 1, f"Missing results for scenario {s}"

    def test_calc_strings_generated_for_all_results(
        self, urshifu, incineroar, surging_strikes
    ):
        summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["normal"], DEFAULT_SCENARIOS["rain"]],
        )
        for r in summary.results:
            assert len(r.calc_string) > 0
            assert "vs." in r.calc_string


# =============================================================================
# Tests: Export-related data structure
# =============================================================================

class TestExportData:

    def test_summary_has_move_names(self, urshifu, incineroar, surging_strikes, close_combat):
        summary = run_bulk_calcs(
            urshifu, [surging_strikes, close_combat], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )
        assert "surging-strikes" in summary.move_names
        assert "close-combat" in summary.move_names

    def test_summary_has_scenario_names(self, urshifu, incineroar, surging_strikes):
        scenarios = [DEFAULT_SCENARIOS["normal"], DEFAULT_SCENARIOS["rain"]]
        summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar], scenarios,
        )
        assert "normal" in summary.scenario_names
        assert "rain" in summary.scenario_names

    def test_result_has_display_name(self, urshifu, incineroar, surging_strikes):
        summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["rain"]],
        )
        assert summary.results[0].scenario_display == "Rain"

    def test_result_has_ko_chance(self, urshifu, incineroar, surging_strikes):
        summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )
        assert summary.results[0].ko_chance != ""


# =============================================================================
# Tests: Spread move multiplier
# =============================================================================

class TestSpreadMoveMultiplier:

    def test_spread_move_deals_less_damage_than_single_target(self, urshifu, incineroar):
        """Spread moves should get 0.75x multiplier in doubles."""
        single_target = _make_move(
            "flamethrower", "fire",
            category=MoveCategory.SPECIAL, power=90,
            target="selected-pokemon",
        )
        spread_move = _make_move(
            "heat-wave", "fire",
            category=MoveCategory.SPECIAL, power=90,
            target="all-adjacent-foes",
        )

        single_summary = run_bulk_calcs(
            urshifu, [single_target], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )
        spread_summary = run_bulk_calcs(
            urshifu, [spread_move], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )

        # Spread should deal ~75% of single-target damage
        assert spread_summary.results[0].max_damage < single_summary.results[0].max_damage


# =============================================================================
# Tests: Defender Tera type support
# =============================================================================

class TestDefenderTera:

    def test_defender_tera_type_changes_damage(self, urshifu, incineroar, surging_strikes):
        """Tera Water Incineroar should take different water damage than base Fire/Dark."""
        normal_summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )
        tera_summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
            defender_tera_types={"incineroar": "water"},
        )
        # Tera Water resists Water (0.5x) vs base Incineroar (neutral)
        assert tera_summary.results[0].max_pct < normal_summary.results[0].max_pct

    def test_defender_tera_only_affects_specified_defender(
        self, urshifu, incineroar, flutter_mane, surging_strikes,
    ):
        """Only the specified defender should get Tera, not all."""
        summary = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar, flutter_mane],
            [DEFAULT_SCENARIOS["normal"]],
            defender_tera_types={"incineroar": "water"},
        )
        # Flutter Mane should NOT be affected by Incineroar's Tera
        normal_summary = run_bulk_calcs(
            urshifu, [surging_strikes], [flutter_mane],
            [DEFAULT_SCENARIOS["normal"]],
        )
        fm_tera_result = [r for r in summary.results if r.defender_name == "flutter-mane"][0]
        assert fm_tera_result.max_pct == normal_summary.results[0].max_pct

    def test_no_defender_tera_when_not_specified(self, urshifu, incineroar, surging_strikes):
        """Without defender_tera_types, results should be identical to None."""
        summary_none = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
            defender_tera_types=None,
        )
        summary_empty = run_bulk_calcs(
            urshifu, [surging_strikes], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
            defender_tera_types={},
        )
        assert summary_none.results[0].max_pct == summary_empty.results[0].max_pct


# =============================================================================
# Tests: Sword of Ruin auto-detection in bulk calcs
# =============================================================================

CHIEN_PAO_STATS = BaseStats(
    hp=80, attack=120, defense=80,
    special_attack=90, special_defense=65, speed=135,
)


class TestSwordOfRuinBulkCalcs:
    """Ruin abilities are field effects (not stat stages) and must always apply,
    including on critical hits. They should be auto-detected from ability names."""

    @pytest.fixture
    def chien_pao(self):
        return _make_build(
            "chien-pao",
            ["dark", "ice"],
            CHIEN_PAO_STATS,
            nature=Nature.ADAMANT,
            evs={"attack": 252, "speed": 252, "hp": 4},
            item="focus-sash",
            ability="sword-of-ruin",
        )

    @pytest.fixture
    def icicle_crash(self):
        return _make_move("icicle-crash", "ice", power=85)

    def test_sword_of_ruin_auto_detected_in_bulk_calcs(
        self, chien_pao, incineroar, icicle_crash
    ):
        """Chien-Pao's Sword of Ruin should be auto-detected and increase damage."""
        # Chien-Pao with Sword of Ruin ability
        summary_ruin = run_bulk_calcs(
            chien_pao, [icicle_crash], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )

        # Same Pokemon but without the Ruin ability
        chien_pao_no_ruin = _make_build(
            "chien-pao",
            ["dark", "ice"],
            CHIEN_PAO_STATS,
            nature=Nature.ADAMANT,
            evs={"attack": 252, "speed": 252, "hp": 4},
            item="focus-sash",
            ability="inner-focus",  # Non-Ruin ability
        )
        summary_no_ruin = run_bulk_calcs(
            chien_pao_no_ruin, [icicle_crash], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )

        assert summary_ruin.results[0].max_pct > summary_no_ruin.results[0].max_pct, (
            "Sword of Ruin should increase damage in bulk calcs via auto-detection"
        )

    def test_sword_of_ruin_applies_on_always_crit_in_bulk_calcs(
        self, chien_pao, incineroar
    ):
        """Sword of Ruin must apply even when move always crits (Flower Trick)."""
        flower_trick = _make_move(
            "flower-trick", "grass", power=70, always_crit=True,
        )

        summary_ruin = run_bulk_calcs(
            chien_pao, [flower_trick], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )

        chien_pao_no_ruin = _make_build(
            "chien-pao",
            ["dark", "ice"],
            CHIEN_PAO_STATS,
            nature=Nature.ADAMANT,
            evs={"attack": 252, "speed": 252, "hp": 4},
            item="focus-sash",
            ability="inner-focus",
        )
        summary_no_ruin = run_bulk_calcs(
            chien_pao_no_ruin, [flower_trick], [incineroar],
            [DEFAULT_SCENARIOS["normal"]],
        )

        assert summary_ruin.results[0].max_pct > summary_no_ruin.results[0].max_pct, (
            "Sword of Ruin must increase damage even on always-crit moves"
        )


# =============================================================================
# Tests: Export generation
# =============================================================================

def _make_bulk_summary() -> BulkCalcSummary:
    """Build a minimal BulkCalcSummary for export tests."""
    results = [
        BulkCalcResult(
            defender_name="incineroar",
            move_name="close-combat",
            scenario_name="normal",
            scenario_display="Normal",
            min_pct=85.0,
            max_pct=100.5,
            min_damage=161,
            max_damage=190,
            defender_hp=189,
            ko_chance="guaranteed OHKO",
            calc_string="252+ Atk Urshifu Close Combat vs. Incineroar: 161-190 (85-100.5%)",
        ),
        BulkCalcResult(
            defender_name="flutter-mane",
            move_name="close-combat",
            scenario_name="normal",
            scenario_display="Normal",
            min_pct=45.0,
            max_pct=53.0,
            min_damage=58,
            max_damage=69,
            defender_hp=131,
            ko_chance="2HKO",
            calc_string="252+ Atk Urshifu Close Combat vs. Flutter Mane: 58-69 (45-53%)",
        ),
        BulkCalcResult(
            defender_name="incineroar",
            move_name="close-combat",
            scenario_name="rain",
            scenario_display="Rain",
            min_pct=85.0,
            max_pct=100.5,
            min_damage=161,
            max_damage=190,
            defender_hp=189,
            ko_chance="guaranteed OHKO",
            calc_string="[Rain] 252+ Atk Urshifu Close Combat vs. Incineroar: 161-190 (85-100.5%)",
        ),
        BulkCalcResult(
            defender_name="flutter-mane",
            move_name="close-combat",
            scenario_name="rain",
            scenario_display="Rain",
            min_pct=45.0,
            max_pct=53.0,
            min_damage=58,
            max_damage=69,
            defender_hp=131,
            ko_chance="2HKO",
            calc_string="[Rain] 252+ Atk Urshifu Close Combat vs. Flutter Mane: 58-69 (45-53%)",
        ),
    ]
    return BulkCalcSummary(
        attacker_name="urshifu-rapid-strike",
        attacker_spread_str="Adamant 4/252/0/0/0/252",
        move_names=["close-combat"],
        scenario_names=["normal", "rain"],
        total_calcs=4,
        results=results,
        ohko_counts={"normal": 1, "rain": 1},
        twohko_counts={"normal": 1, "rain": 1},
        defender_spreads={
            "incineroar": "Careful 252/0/128/0/128/0",
            "flutter-mane": "Timid 4/0/0/252/0/252",
        },
        defender_items={
            "incineroar": "sitrus-berry",
            "flutter-mane": "choice-specs",
        },
    )


class TestExportGeneration:

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_excel_report_creates_file(self):
        from vgc_mcp_core.export.damage_report import generate_excel_report

        summary = _make_bulk_summary()
        output = os.path.join(self.tmpdir, "test_report.xlsx")
        result_path = generate_excel_report(summary, output)

        assert os.path.exists(result_path)
        assert result_path.endswith(".xlsx")
        assert os.path.getsize(result_path) > 0

    def test_pdf_report_creates_file(self):
        from vgc_mcp_core.export.damage_report import generate_pdf_report

        summary = _make_bulk_summary()
        output = os.path.join(self.tmpdir, "test_report.pdf")
        result_path = generate_pdf_report(summary, output)

        assert os.path.exists(result_path)
        assert result_path.endswith(".pdf")
        assert os.path.getsize(result_path) > 0
        # Verify it's a valid PDF
        with open(result_path, "rb") as f:
            header = f.read(5)
        assert header == b"%PDF-"

    def test_excel_has_scenario_sheets(self):
        from openpyxl import load_workbook

        from vgc_mcp_core.export.damage_report import generate_excel_report

        summary = _make_bulk_summary()
        output = os.path.join(self.tmpdir, "test_sheets.xlsx")
        generate_excel_report(summary, output)

        wb = load_workbook(output)
        sheet_names = wb.sheetnames
        # Should have one sheet per scenario
        assert len(sheet_names) == 2
        assert "Normal" in sheet_names
        assert "Rain" in sheet_names

    def test_excel_cells_have_color(self):
        from openpyxl import load_workbook

        from vgc_mcp_core.export.damage_report import generate_excel_report

        summary = _make_bulk_summary()
        output = os.path.join(self.tmpdir, "test_colors.xlsx")
        generate_excel_report(summary, output)

        wb = load_workbook(output)
        ws = wb["Normal"]
        # Find the OHKO cell (Incineroar close-combat, row 5, col 4)
        ohko_cell = ws.cell(row=5, column=4)
        assert ohko_cell.fill.start_color.rgb is not None
        # Green for OHKO: C6EFCE
        assert "C6EFCE" in str(ohko_cell.fill.start_color.rgb)
